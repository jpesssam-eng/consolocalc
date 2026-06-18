"""
ConsoloCalc - Modulo de exportacao Excel/CSV
==============================================
Funcoes para exportar os resultados do dimensionamento em formatos
consumiveis por Excel, Power BI, Google Sheets e outras ferramentas
de analise de dados.

Autor: Junior Pessoa da Silva
TCC - Engenharia Civil - UNINTER - 2026
"""

from io import BytesIO, StringIO
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ============================================================
# Estilos (paleta consistente com o PDF)
# ============================================================
COR_PRIMARIA = "1A4F7A"
COR_FUNDO_LINHA = "F0F0F0"
COR_OK = "2E7D32"
COR_ERRO = "C62828"

FONTE_TITULO = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONTE_CABECALHO = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONTE_NORMAL = Font(name="Calibri", size=11)
FONTE_NEGRITO = Font(name="Calibri", size=11, bold=True)

FILL_TITULO = PatternFill(start_color=COR_PRIMARIA, end_color=COR_PRIMARIA,
                          fill_type="solid")
FILL_CABECALHO = PatternFill(start_color=COR_PRIMARIA, end_color=COR_PRIMARIA,
                             fill_type="solid")
FILL_ZEBRA = PatternFill(start_color=COR_FUNDO_LINHA, end_color=COR_FUNDO_LINHA,
                         fill_type="solid")

BORDA_FINA = Border(
    left=Side(style="thin", color="888888"),
    right=Side(style="thin", color="888888"),
    top=Side(style="thin", color="888888"),
    bottom=Side(style="thin", color="888888"),
)

CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)
ESQUERDA = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _aplicar_estilo_cabecalho(celula):
    """Aplica estilo padrao de cabecalho a uma celula."""
    celula.font = FONTE_CABECALHO
    celula.fill = FILL_CABECALHO
    celula.alignment = CENTRO
    celula.border = BORDA_FINA


def _aplicar_zebra(linha_celulas, indice_linha):
    """Aplica fundo zebra (alternado) e bordas finas."""
    for celula in linha_celulas:
        celula.border = BORDA_FINA
        if indice_linha % 2 == 0:
            celula.fill = FILL_ZEBRA


def _ajustar_larguras(ws, larguras):
    """Aplica larguras de coluna por indice."""
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura


def gerar_excel(dados, resultado) -> bytes:
    """
    Gera planilha Excel (.xlsx) com 3 abas:
        - Dados de Entrada
        - Resultados
        - Memorial (texto consolidado)
    Retorna os bytes do arquivo, prontos para download.
    """
    wb = openpyxl.Workbook()

    # ============================================================
    # ABA 1 - Dados de Entrada
    # ============================================================
    ws1 = wb.active
    ws1.title = "Dados de Entrada"

    # Titulo
    ws1["A1"] = "DADOS DE ENTRADA — CONSOLOCALC"
    ws1["A1"].font = FONTE_TITULO
    ws1["A1"].fill = FILL_TITULO
    ws1["A1"].alignment = CENTRO
    ws1.merge_cells("A1:C1")
    ws1.row_dimensions[1].height = 28

    # Data
    ws1["A2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y as %H:%M')}"
    ws1["A2"].font = Font(italic=True, color="666666")
    ws1.merge_cells("A2:C2")

    # Cabecalho
    headers = ["Parametro", "Valor", "Unidade"]
    for col, valor in enumerate(headers, start=1):
        c = ws1.cell(row=4, column=col, value=valor)
        _aplicar_estilo_cabecalho(c)

    # Linhas
    linhas_dados = [
        ("fck (resistencia do concreto)", dados.fck, "MPa"),
        ("fyk (resistencia do aco)", dados.fyk, "MPa"),
        ("a (distancia da carga a face do pilar)", dados.a, "cm"),
        ("h (altura do consolo)", dados.h, "cm"),
        ("b (largura do consolo)", dados.b, "cm"),
        ("d' (cobrimento ate CG da armadura)", dados.d_linha, "cm"),
        ("Vk (forca vertical caracteristica)", dados.Vk, "kN"),
        ("Hk (forca horizontal caracteristica)", dados.Hk, "kN"),
        ("Impedimento horizontal",
         "Sim" if dados.impedimento_horizontal else "Nao", "-"),
    ]

    for i, (param, valor, unidade) in enumerate(linhas_dados):
        linha = i + 5
        ws1.cell(row=linha, column=1, value=param).font = FONTE_NORMAL
        ws1.cell(row=linha, column=1).alignment = ESQUERDA
        ws1.cell(row=linha, column=2, value=valor).font = FONTE_NEGRITO
        ws1.cell(row=linha, column=2).alignment = CENTRO
        ws1.cell(row=linha, column=3, value=unidade).font = FONTE_NORMAL
        ws1.cell(row=linha, column=3).alignment = CENTRO
        _aplicar_zebra(
            [ws1.cell(row=linha, column=c) for c in range(1, 4)], i + 1
        )

    _ajustar_larguras(ws1, [40, 15, 12])

    # ============================================================
    # ABA 2 - Resultados
    # ============================================================
    ws2 = wb.create_sheet("Resultados")

    ws2["A1"] = "RESULTADOS DO DIMENSIONAMENTO"
    ws2["A1"].font = FONTE_TITULO
    ws2["A1"].fill = FILL_TITULO
    ws2["A1"].alignment = CENTRO
    ws2.merge_cells("A1:C1")
    ws2.row_dimensions[1].height = 28

    # Cabecalho
    for col, valor in enumerate(["Variavel", "Valor", "Unidade"], start=1):
        c = ws2.cell(row=3, column=col, value=valor)
        _aplicar_estilo_cabecalho(c)

    biela = resultado["biela"]
    linhas_result = [
        ("Tipo do consolo", f"consolo {resultado['tipo']}", "-"),
        ("Relacao a/d", round(resultado["razao_ad"], 3), "-"),
        ("Altura util d", round(resultado["d"], 2), "cm"),
        ("Vd (forca vertical de calculo)", round(resultado["Vd"], 2), "kN"),
        ("Hd (forca horizontal de calculo)", round(resultado["Hd"], 2), "kN"),
        ("Excentricidade e", round(resultado["e"], 3), "cm"),
        ("Braco de alavanca z", round(resultado["z"], 2), "cm"),
        ("Angulo da biela theta", round(resultado["theta_deg"], 2), "graus"),
        ("Rsd (forca no tirante)", round(resultado["Rsd"], 2), "kN"),
        ("Fc (forca na biela)", round(resultado["Fc"], 2), "kN"),
        ("As tirante", round(resultado["As_tirante"], 3), "cm2"),
        ("As costura", round(resultado["As_costura"], 3), "cm2"),
        ("Sigma atuante na biela", round(biela["sigma_atuante_MPa"], 2), "MPa"),
        ("Sigma limite na biela", round(biela["sigma_limite_MPa"], 2), "MPa"),
        ("Aproveitamento da biela", round(biela["razao"] * 100, 1), "%"),
        ("Status da biela", "OK" if biela["ok"] else "ESMAGADA", "-"),
    ]

    for i, (var, valor, unidade) in enumerate(linhas_result):
        linha = i + 4
        ws2.cell(row=linha, column=1, value=var).font = FONTE_NORMAL
        ws2.cell(row=linha, column=1).alignment = ESQUERDA
        ws2.cell(row=linha, column=2, value=valor).font = FONTE_NEGRITO
        ws2.cell(row=linha, column=2).alignment = CENTRO
        ws2.cell(row=linha, column=3, value=unidade).font = FONTE_NORMAL
        ws2.cell(row=linha, column=3).alignment = CENTRO
        _aplicar_zebra(
            [ws2.cell(row=linha, column=c) for c in range(1, 4)], i + 1
        )

    # Destacar status da biela
    linha_status = 4 + len(linhas_result) - 1
    cor_status = COR_OK if biela["ok"] else COR_ERRO
    ws2.cell(row=linha_status, column=2).font = Font(
        name="Calibri", size=11, bold=True, color="FFFFFF"
    )
    ws2.cell(row=linha_status, column=2).fill = PatternFill(
        start_color=cor_status, end_color=cor_status, fill_type="solid"
    )

    _ajustar_larguras(ws2, [38, 18, 10])

    # ============================================================
    # ABA 3 - Memorial (texto consolidado para citacao)
    # ============================================================
    ws3 = wb.create_sheet("Memorial")

    ws3["A1"] = "MEMORIAL TEXTUAL — PRONTO PARA CITACAO EM RELATORIOS"
    ws3["A1"].font = FONTE_TITULO
    ws3["A1"].fill = FILL_TITULO
    ws3["A1"].alignment = CENTRO
    ws3.merge_cells("A1:B1")
    ws3.row_dimensions[1].height = 28

    blocos_texto = [
        ("Classificacao",
         f"O consolo apresenta relacao a/d = {resultado['razao_ad']:.3f}, "
         f"sendo classificado como consolo {resultado['tipo']}."),
        ("Esforcos de calculo",
         f"Vd = 1,4 * Vk = {resultado['Vd']:.2f} kN. "
         f"Hd = {resultado['Hd']:.2f} kN. "
         f"Excentricidade e = {resultado['e']:.3f} cm."),
        ("Modelo de bielas e tirantes",
         f"Braco z = 0,85 d = {resultado['z']:.2f} cm. "
         f"Angulo da biela theta = {resultado['theta_deg']:.2f} graus. "
         f"Forca de tracao no tirante Rsd = {resultado['Rsd']:.2f} kN. "
         f"Forca de compressao na biela Fc = {resultado['Fc']:.2f} kN."),
        ("Armaduras",
         f"As tirante = {resultado['As_tirante']:.2f} cm2. "
         f"As costura = {resultado['As_costura']:.2f} cm2."),
        ("Verificacao da biela",
         f"Tensao atuante = {biela['sigma_atuante_MPa']:.2f} MPa, "
         f"tensao limite = {biela['sigma_limite_MPa']:.2f} MPa "
         f"(aproveitamento de {biela['razao']*100:.1f}%). "
         f"Status: {'biela atende ao limite' if biela['ok'] else 'biela esmagada'}."),
    ]

    for i, (titulo, texto) in enumerate(blocos_texto):
        linha = i * 3 + 3
        ws3.cell(row=linha, column=1, value=titulo).font = Font(
            name="Calibri", size=12, bold=True, color=COR_PRIMARIA
        )
        ws3.cell(row=linha + 1, column=1, value=texto).font = FONTE_NORMAL
        ws3.cell(row=linha + 1, column=1).alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        ws3.merge_cells(start_row=linha + 1, start_column=1,
                        end_row=linha + 1, end_column=2)
        ws3.row_dimensions[linha + 1].height = 45

    _ajustar_larguras(ws3, [60, 30])

    # ============================================================
    # Salvar em bytes
    # ============================================================
    buffer = BytesIO()
    wb.save(buffer)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def gerar_csv(dados, resultado) -> str:
    """
    Gera CSV simples com todos os resultados em formato chave-valor.
    Util para Power BI, Google Sheets e analises rapidas.
    """
    biela = resultado["biela"]

    linhas = [
        "categoria,parametro,valor,unidade",
        f"entrada,fck,{dados.fck},MPa",
        f"entrada,fyk,{dados.fyk},MPa",
        f"entrada,a,{dados.a},cm",
        f"entrada,h,{dados.h},cm",
        f"entrada,b,{dados.b},cm",
        f"entrada,d_linha,{dados.d_linha},cm",
        f"entrada,Vk,{dados.Vk},kN",
        f"entrada,Hk,{dados.Hk},kN",
        f"entrada,impedimento_horizontal,{int(dados.impedimento_horizontal)},bool",
        f"resultado,tipo,consolo {resultado['tipo']},texto",
        f"resultado,razao_ad,{resultado['razao_ad']:.4f},adimensional",
        f"resultado,d,{resultado['d']:.2f},cm",
        f"resultado,Vd,{resultado['Vd']:.2f},kN",
        f"resultado,Hd,{resultado['Hd']:.2f},kN",
        f"resultado,e,{resultado['e']:.4f},cm",
        f"resultado,z,{resultado['z']:.2f},cm",
        f"resultado,theta,{resultado['theta_deg']:.2f},graus",
        f"resultado,Rsd,{resultado['Rsd']:.2f},kN",
        f"resultado,Fc,{resultado['Fc']:.2f},kN",
        f"resultado,As_tirante,{resultado['As_tirante']:.3f},cm2",
        f"resultado,As_costura,{resultado['As_costura']:.3f},cm2",
        f"verificacao,sigma_atuante,{biela['sigma_atuante_MPa']:.2f},MPa",
        f"verificacao,sigma_limite,{biela['sigma_limite_MPa']:.2f},MPa",
        f"verificacao,aproveitamento_biela,{biela['razao']*100:.2f},%",
        f"verificacao,status_biela,{'OK' if biela['ok'] else 'ESMAGADA'},texto",
    ]

    return "\n".join(linhas)
